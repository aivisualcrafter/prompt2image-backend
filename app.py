import os
from flask import Flask, request, render_template, send_from_directory
from generate_images import generate_image
from openpyxl import load_workbook

app = Flask(__name__)
UPLOAD_FOLDER = 'static'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['excel']
        if file:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, desc, prompt = row
                generate_image(prompt, name, desc)
            return "Images generated successfully!"
    return render_template('index.html')

@app.route('/static/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)